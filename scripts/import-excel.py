"""
Import script: Boekhouding 2026.xlsx → Supabase

Vereisten:
  pip install openpyxl requests

Gebruik:
  1. Zet je SUPABASE_URL en SUPABASE_KEY hieronder (of als omgevingsvariabele)
  2. Run: python3 scripts/import-excel.py

Het script importeert:
- Alle verkopen uit het 'Verkoop' tabblad
- Alle inkopen uit het 'Voorraad' tabblad
"""

import os
import json
import requests
import openpyxl
from datetime import datetime, date

# === CONFIGURATIE ===
SUPABASE_URL = os.environ.get("NEXT_PUBLIC_SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("NEXT_PUBLIC_SUPABASE_ANON_KEY", "")
EXCEL_BESTAND = "../Boekhouding 2026.xlsx"

if not SUPABASE_URL or not SUPABASE_KEY:
    print("❌ Zet NEXT_PUBLIC_SUPABASE_URL en NEXT_PUBLIC_SUPABASE_ANON_KEY als omgevingsvariabelen")
    print("   Voorbeeld: NEXT_PUBLIC_SUPABASE_URL=https://xxx.supabase.co NEXT_PUBLIC_SUPABASE_ANON_KEY=eyJ... python3 scripts/import-excel.py")
    exit(1)

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=minimal",
}

# Mapping van oude accountnamen naar de 2 nieuwe accounts
# Pas dit aan als een naam verkeerd is gemapt
ACCOUNT_MAPPING = {
    "jesuslata - laptop": "1-jesuslata",
    "jesuslata": "1-jesuslata",
    "disteltr - moto": "2-disteltr",
    "disteltr": "2-disteltr",
    "moto g54": "2-disteltr",
    "17 - tristanjansse": "1-jesuslata",
    "tristanjansse": "1-jesuslata",
    "-": "1-jesuslata",
}

def normaliseer_account(account: str) -> str:
    """Normaliseer accountnaam (case-insensitief) naar de correcte waarde."""
    if not account:
        return "1-jesuslata"
    mapped = ACCOUNT_MAPPING.get(account.strip().lower())
    if mapped:
        return mapped
    # Als de naam al een van de twee geldige accounts is, gebruik die
    if account.strip() in ("1-jesuslata", "2-disteltr"):
        return account.strip()
    # Onbekende naam: bewaar origineel zodat je het later kan controleren
    print(f"  ⚠ Onbekend account '{account}', bewaard als origineel")
    return account.strip()

GELDIGE_STATUSSEN = [
    "Afgerond (geld binnen)",
    "Verkocht - Nog niet verzonden",
    "Onderweg",
    "Retour",
    "Retour ontvangen",
    "Verlies",
    "Probleem",
]

# Normaliseer schrijffouten uit het Excel bestand
STATUS_MAPPING = {
    "verkocht - nog niet verzonden": "Verkocht - Nog niet verzonden",
    "afgerond (geld binnen)": "Afgerond (geld binnen)",
    "onderweg": "Onderweg",
    "retour": "Retour",
    "retour ontvangen": "Retour ontvangen",
    "verlies": "Verlies",
    "probleem": "Probleem",
}

def normaliseer_status(status: str) -> str:
    """Normaliseer status (case-insensitief) naar de correcte waarde."""
    return STATUS_MAPPING.get(status.strip().lower(), status.strip())

def datum_str(waarde):
    """Zet Excel datum om naar YYYY-MM-DD string."""
    if waarde is None:
        return None
    if isinstance(waarde, (datetime, date)):
        return waarde.strftime("%Y-%m-%d")
    if isinstance(waarde, str):
        waarde = waarde.strip()
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(waarde, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return None

def batch_insert(tabel, rijen, batch_grootte=100):
    """Voeg rijen in batches in via Supabase REST API."""
    totaal = len(rijen)
    ingevoegd = 0
    fouten = 0

    for i in range(0, totaal, batch_grootte):
        batch = rijen[i : i + batch_grootte]
        url = f"{SUPABASE_URL}/rest/v1/{tabel}"
        res = requests.post(url, headers=HEADERS, json=batch)
        if res.status_code in (200, 201):
            ingevoegd += len(batch)
            print(f"  ✓ {ingevoegd}/{totaal}")
        else:
            fouten += len(batch)
            print(f"  ✗ Fout bij batch {i//batch_grootte + 1}: {res.status_code} {res.text[:200]}")

    return ingevoegd, fouten

def import_verkopen(ws):
    """Importeer het Verkoop tabblad."""
    print("\n📋 Verkopen importeren...")
    rijen = []
    max_rijen = ws.max_row

    for row_idx in range(2, max_rijen + 1):
        verkoopdatum = datum_str(ws.cell(row=row_idx, column=2).value)  # B
        product = ws.cell(row=row_idx, column=3).value  # C
        naam_koper = ws.cell(row=row_idx, column=4).value  # D
        verkoopprijs = ws.cell(row=row_idx, column=6).value  # F
        status = ws.cell(row=row_idx, column=7).value  # G
        account = ws.cell(row=row_idx, column=8).value  # H

        # Sla lege rijen over
        if not verkoopdatum or not product or not naam_koper:
            continue

        # Normaliseer en valideer status
        if status:
            status = normaliseer_status(str(status))
        if not status or status not in GELDIGE_STATUSSEN:
            print(f"  ⚠ Rij {row_idx}: onbekende status '{status}', overgeslagen")
            continue

        # Verkoopprijs
        try:
            verkoopprijs = float(verkoopprijs) if verkoopprijs is not None else 0.0
        except (TypeError, ValueError):
            verkoopprijs = 0.0

        rijen.append({
            "verkoopdatum": verkoopdatum,
            "product": str(product).strip(),
            "naam_koper": str(naam_koper).strip(),
            "verkoopprijs": verkoopprijs,
            "status": str(status).strip(),
            "account": normaliseer_account(str(account)) if account else "1-jesuslata",
        })

    print(f"  Gevonden: {len(rijen)} verkopen")
    return batch_insert("verkopen", rijen)

def import_inkopen(ws):
    """Importeer bestellingen uit het Voorraad tabblad (rij 22+)."""
    print("\n📦 Inkopen importeren...")
    rijen = []
    max_rijen = ws.max_row

    # Bestellingen starten na rij 21 (header rij 21)
    for row_idx in range(22, max_rijen + 1):
        besteldatum = datum_str(ws.cell(row=row_idx, column=1).value)  # A
        product = ws.cell(row=row_idx, column=2).value  # B
        aantal = ws.cell(row=row_idx, column=3).value  # C
        status = ws.cell(row=row_idx, column=4).value  # D
        totale_aankoopprijs = ws.cell(row=row_idx, column=5).value  # E
        prijs_per_stuk = ws.cell(row=row_idx, column=6).value  # F

        # Sla lege rijen over
        if not besteldatum or not product or not aantal:
            continue

        try:
            aantal = int(float(str(aantal)))
        except (TypeError, ValueError):
            continue

        try:
            totaal = float(totale_aankoopprijs) if totale_aankoopprijs is not None else 0.0
        except (TypeError, ValueError):
            totaal = 0.0

        try:
            pps = float(prijs_per_stuk) if prijs_per_stuk is not None else 0.0
        except (TypeError, ValueError):
            pps = totaal / aantal if aantal > 0 else 0.0

        # Valideer status
        status_str = str(status).strip() if status else "In Huis"
        if status_str not in ("In Huis", "Onderweg"):
            status_str = "In Huis"

        rijen.append({
            "besteldatum": besteldatum,
            "product": str(product).strip(),
            "aantal": aantal,
            "status": status_str,
            "totale_aankoopprijs": totaal,
            "prijs_per_stuk": pps,
        })

    print(f"  Gevonden: {len(rijen)} inkoopregels")
    return batch_insert("inkopen", rijen)

def main():
    print("🚀 Vinted Boekhouding — Excel Import")
    print(f"📂 Bestand: {EXCEL_BESTAND}")
    print(f"🔗 Supabase: {SUPABASE_URL}")

    # Laad Excel
    try:
        wb = openpyxl.load_workbook(EXCEL_BESTAND, data_only=True)
    except FileNotFoundError:
        print(f"\n❌ Bestand niet gevonden: {EXCEL_BESTAND}")
        print("   Zorg dat je het script vanuit de vinted-dashboard map uitvoert")
        exit(1)

    print(f"\n📊 Tabbladen gevonden: {wb.sheetnames}")

    # Test Supabase verbinding
    test_url = f"{SUPABASE_URL}/rest/v1/verkopen?limit=1"
    test_res = requests.get(test_url, headers=HEADERS)
    if test_res.status_code != 200:
        print(f"\n❌ Supabase verbinding mislukt: {test_res.status_code}")
        print("   Controleer je URL en API key, en of de tabellen aangemaakt zijn")
        exit(1)
    print("✅ Supabase verbinding werkt")

    totaal_v_ok = totaal_v_fout = 0
    totaal_i_ok = totaal_i_fout = 0

    # Importeer Verkoop
    if "Verkoop" in wb.sheetnames:
        ok, fout = import_verkopen(wb["Verkoop"])
        totaal_v_ok, totaal_v_fout = ok, fout
    else:
        print("⚠ Tabblad 'Verkoop' niet gevonden")

    # Importeer Voorraad/inkopen
    if "Voorraad" in wb.sheetnames:
        ok, fout = import_inkopen(wb["Voorraad"])
        totaal_i_ok, totaal_i_fout = ok, fout
    else:
        print("⚠ Tabblad 'Voorraad' niet gevonden")

    print("\n" + "="*40)
    print(f"✅ Verkopen ingevoegd: {totaal_v_ok}")
    print(f"✅ Inkopen ingevoegd:  {totaal_i_ok}")
    if totaal_v_fout or totaal_i_fout:
        print(f"❌ Fouten: {totaal_v_fout + totaal_i_fout}")
    print("🎉 Import voltooid!")

if __name__ == "__main__":
    main()
